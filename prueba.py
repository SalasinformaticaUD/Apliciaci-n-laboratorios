from flask import Flask, Response, send_file
from flask import jsonify
from flask import request
from flask import session
from flask_restful import Resource, Api
from flask_pymongo import PyMongo
from flask_cors import CORS
import datetime
import json
from bson import json_util, ObjectId
from flask_mail import Mail,  Message


from openpyxl.utils import get_column_letter
from openpyxl import Workbook, cell
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, Reference, Series

import gspread
from oauth2client.service_account import ServiceAccountCredentials


app = Flask(__name__)
CORS(app)


app.secret_key = 'clavedelaaplicacion'
api = Api(app)

mongo = PyMongo(app, 'mongodb://localhost:27017/Aplicacion')
# print(dir(mongo.db.list_collection_names()))

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USERNAME'] = 'rjchiaj@correo.udistrital.edu.co'
app.config['MAIL_PASSWORD'] = 'At3QM$por/'
app.config['MAIL_USE_TLS'] = True

mail = Mail(app)

class Usuario(Resource):
    status = 0
    datos = {}
    mensaje = ""
    token = '79548af95a6f0d909441d77c3814a3f36030dc317d05a9204dfd8172b7f51b6'

    def login(self):
        user = mongo.db.Usuarios
        usuario = request.json['usuario']
        contrasena = request.json['contrasena']
        tipo= ""
        pregunta = user.find_one(
            {"Código_usuario": usuario, "Contraseña": contrasena})  
        if pregunta:
            tipo = pregunta['Tipo']
            print(tipo)
            mensaje = "1"
            self.status = 1
            if tipo == "Laboratorista":
               self.datos = {'token': self.token, 'addr':"HomeLaboratorista", 'kind':"1"}       
            elif tipo == "Administrador":
                self.datos = {'token': self.token, 'addr':"HomeAdmin", 'kind':"2"}
            elif tipo == "Estudiante":
                self.datos = {'token': self.token, 'addr':"reservaestudiante", 'kind':"3"}
        else:  
            mensaje = "0"
            print(mensaje)
            self.status = 1    
        return mensaje

    def isLoggedIn(self):
        return self.token

    def cerrarSesion(self):
        session.pop('tipoUsuario', None)
        return "1"

    def registrar(self):
        mensaje = "0"
        user = mongo.db.Usuarios
        usuario = request.json['usuario']
        identificacion = request.json['identificacion']
        correo = request.json['correo']
        estado = request.json['estado']
        proyecto = request.json['proyecto']
        contraseña = request.json['contraseña']
        tipousuario = request.json['tipousuario']
        pregunta = user.find_one({"Código_usuario": identificacion})
        if pregunta:
            mensaje = "Este Laboratorista ya existe en el sistema"
            self.status = 1
        else:
            nuevoUsuario = user.insert({'Nombre_usuario': usuario, 'Código_usuario': identificacion, 'Correo_usuario': correo,
                                        'Código_Estado': estado, 'Proyecto_Curricular': proyecto, 'Contraseña': contraseña, 'Tipo': tipousuario})
            mensaje = "Usuario Registrado correctamente"
            self.status = 1
        return mensaje

    def desactivar(self):
        user = mongo.db.Usuarios
        codigo = request.json['codigo']
        print(codigo)
        estado = "NA"
        mensaje = "Desactivado"
        desactivar = user.update({"Código_usuario": codigo}, {
                                 "$set": {"Código_Estado": estado}})
        if desactivar:
            self.status = 1
            mensaje
        return mensaje

    def editaruser(self):
        user = mongo.db.Usuarios
        codigo = request.json['codigo']
        usuario = request.json['usuario']
        correo = request.json['correo']
        mensaje = "Laboratorista actualizado"
        actualizardatos = user.update({"Código_usuario": codigo}, {"$set": {
                                      "Nombre_usuario": usuario, "Correo_usuario": correo}})
        if actualizardatos:
            self.status = 1
            mensaje
        return mensaje

    def resetpass(self):
        user = mongo.db.Usuarios
        codigo = request.json['codigo']
        contraseña = request.json['codigo']
        mensaje = "Clave restaurada"
        actualizarpass = user.update({"Código_usuario": codigo}, {
                                     "$set": {"Contraseña": contraseña}})
        if actualizarpass:
            self.status = 1
            mensaje
        return mensaje

    def editaruserlab(self):
        user = mongo.db.Usuarios
        codigo = request.json['codigo']
        usuario = request.json['usuario']
        correo = request.json['correo']
        mensaje = "Datos actualizados"
        actualizardatos = user.update({"Código_usuario": codigo}, {
                                      "$set": {"Nombre_usuario": usuario, "Correo_usuario": correo}})
        if actualizardatos:
            self.status = 1
            mensaje
        return mensaje

    def editpass(self):
        user = mongo.db.Usuarios
        codigo = request.json['codigo']
        contraseña = request.json['contraseña']
        nuevacontraseña = request.json['nuevacontraseña']
        mensaje = "Contraseña actualizada"
        actualizarpass = user.update({"$and": [{'Código_usuario': codigo}, {
            'Contraseña': contraseña}]}, {"$set": {"Contraseña": nuevacontraseña}})
        if actualizarpass:
            self.status = 1
            mensaje
        return mensaje

    def editpassadmin(self):
        mensaje = "0"
        user = mongo.db.Usuarios
        codigo = request.json['codigo']
        contraseña = request.json['contraseña']
        nuevacontraseña = request.json['nuevacontraseña']
        pregunta = user.find_one(
            {"$and": [{'Código_usuario': codigo}, {'Contraseña': contraseña}]})
        if pregunta:
            user.update({"$and": [{'Código_usuario': codigo}, {'Contraseña': contraseña}]}, {
                "$set": {"Contraseña": nuevacontraseña}})
            self.status = 1
            mensaje = "Contraseña actualizada"
        else:
            mensaje = "Las contraseñas no corresponden"
            self.status = 2
        return mensaje

    def activar(self):
        user = mongo.db.Usuarios
        codigo = request.json['codigo']
        print(codigo)
        estado = "A"
        mensaje = "Activado"
        activar = user.update({"Código_usuario": codigo}, {
                              "$set": {"Código_Estado": estado}})
        if activar:
            self.status = 1
            mensaje
        return mensaje

    def reserva(self):
        mensaje = ""
        self.status = 2

        # Se toma la fecha de la maquina del usuario (date_user) en formato dia/mes/año y se divide en un vector. Igualmente se toma la hora y los minutos
        date_user = request.json['date_user'].split('/')
        hour_user = request.json['hour_user']
        minutes_user = request.json['minutes_user']
        
        # Con todos los datos de la fecha del usuario se crea un formato de fecha datetime de python
        date_user = datetime.datetime(int(date_user[2]),int(date_user[1]),int(date_user[0]),hour_user,minutes_user,0)
        date_now = datetime.datetime.now()  # Fecha del servidor año-mes-dia hora:min:sec

        # Se toma la diferencia en segundos de las fechas (mayor menos menor o se desborda la resta)
        if(date_user>date_now):
            difference = (date_user - date_now).seconds
        else:
            difference = (date_now - date_user).seconds


        # Si la diferencia no es mayor o menor a 5 minutos (300 segundos) entonces continua, sino, rechaza la peticion
        if(-300<=difference and difference<=300):

            user = mongo.db.Prestamo                    
            hora = request.json['hora']
            codigo = request.json['codigo']
            fecha_adicional = request.json['fecha_adicional']

            # Verifica que no haya otro adicional del usuario en la misma franja horaria
            cruceHora = user.find({'Fecha_Adicional':fecha_adicional, 'Hora':hora, 'Codigo': codigo}).count()

            actual_week = date_now.isocalendar()[1]     # Numero de semana actual ISO
            next_week = actual_week + 1                 # Numero de semana siguiente ISO
            count_actual = 0                            # Contador semana actual
            count_next = 0                              # Contador semana siguiente

            # Numero de semana del adicional
            adic_week = fecha_adicional.split('/')
            adic_week = datetime.datetime(int(adic_week[2]),int(adic_week[1]),int(adic_week[0]))
            adic_week = adic_week.isocalendar()[1]

            # Consulta los adicionales con estado pendiente y con estado aprobado
            reservas = user.find({"$or":[
                                    {"$and": [{'Codigo': codigo}, {'Estado': "PENDIENTE"}]},
                                    {"$and": [{'Codigo': codigo}, {'Estado': "APROBADO"}]}
                                ]})
            
            # Identifica cuantos adicionales coinciden con la semana actual o la semana siguiente
            for item in reservas:
                date_item = item['Fecha_Adicional'].split('/')
                date_item = datetime.datetime(int(date_item[2]),int(date_item[1]),int(date_item[0]))
                week_date_item = date_item.isocalendar()[1]
                count_actual = count_actual+1 if week_date_item==actual_week else count_actual+0
                count_next = count_next+1 if week_date_item==next_week else count_next+0

            if(count_actual==3 and adic_week==actual_week):
                mensaje = "Ya tiene tres laboratorios registrados para esta semana."
            elif (count_next==3 and adic_week==next_week):
                mensaje = "Ya tiene tres laboratorios registrados para la semana siguiente."
            elif (cruceHora>0):
                mensaje = "Ya tiene un adicional pendiente en esta franja horaria."
            else:
                usuario = request.json['usuario']
                dia = request.json['diaSemana']
                sala = request.json['sala']
                banco = request.json['banco']
                practica = request.json['practica']
                elemento = request.json['elemento']
                estado = "PENDIENTE"
                asistencia = "PENDIENTE"                       
                fecha_reserva = date_now.strftime("%d/%m/%Y")
                estadoElemento = [""]*len(elemento)
                placaElemento = [""]*len(elemento)
                monitor = ""
                observacionesGenerales = ""

                user.insert({'Hora': hora, 'Dia': dia, 'Fecha_Adicional': fecha_adicional, 'Fecha_reserva': fecha_reserva,'Codigo': codigo, 'Usuario': usuario, 'Sala': sala, 'practica': practica, 'Banco': banco, 'Elemento': elemento, 'Estado': estado, 'Asistencia': asistencia,'Estado_Elemento': estadoElemento,'Placa_Elemento':placaElemento,'Observaciones_Generales': observacionesGenerales, 'Monitor': monitor})

                mensaje = "Laboratorio registrado correctamente."
                self.status = 1
        else:
            mensaje = "Ocurrio un error de sincronización con la fecha y hora del sistema. Revise su hora y fecha actual."

        return mensaje

    def consultabanco(self):
        mensaje = "0"
        user = mongo.db.Prestamo
        hora = request.json['hora']
        fecha_adicional = request.json['fecha_adicional']
        sala = request.json['sala']
        banco = request.json['banco']
        estado = "PENDIENTE"
        estado2 = "APROBADO"
        pregunta = user.find_one({"$and": [{'Sala': sala}, {'Hora': hora}, {
                                 'Fecha_Adicional': fecha_adicional}, {'Banco': banco},{'Estado':estado}]})
        pregunta2 = user.find_one({"$and": [{'Sala': sala}, {'Hora': hora}, {
                                 'Fecha_Adicional': fecha_adicional}, {'Banco': banco},{'Estado':estado2}]})                                 
                                
        if pregunta2 != None or pregunta != None:
            self.status = 1
            mensaje = "banco ocupado"
        return mensaje

    def buscarreserva(self):
        user = mongo.db.Prestamo
        codigo = request.json['codigo']
        output = []
        self.status = 1
        mensaje = "Encontrado"
        for u in user.find({"Codigo": codigo}):
        	#'asistencia': u['Asistencia'], 'estadoElemento': u['Estado_Elemento'],'placaElemento': u['Placa_Elemento'], 'observacionesGenerales': u['Observaciones_Generales'],
            output.append({'hora': u['Hora'], 'dia': u['Dia'], 'fecha_adicional': u['Fecha_Adicional'],'practica': u['practica'],'asistencia': u['Asistencia'], 'estadoElemento': u['Estado_Elemento'],'placaElemento': u['Placa_Elemento'], 'observacionesGenerales': u['Observaciones_Generales'],
                           'fecha_reserva': u['Fecha_reserva'], 'sala': u['Sala'], 'banco': u['Banco'], 'elemento': u['Elemento'],'usuario': u['Usuario'], 'estado': u['Estado']})
        return {'status': self.status, 'mensaje': mensaje, 'data': output}
    
    def getReservasLaboratorio(self):
        user = mongo.db.Prestamo

        # Agregar campos nuevos a las bases de datos de Mantenimiento

        # mant = mongo.db.Mantenimiento
        # campo = 'Observaciones'
        # for item in mant.find():
        #     if campo in item.keys():
        #         print(item[campo] + " - " +  item['Placa'])
        #     else:
        #         mant.update({'_id': item['_id']}, {"$set": {campo: ""}})        

        fecha = request.json['fecha_laboratorio']
        sala = request.json['sala_laboratorio']
        hora = request.json['hora_laboratorio']
        output = []        
        registros = user.find({'Fecha_Adicional': fecha, 'Hora': hora, 'Sala': sala})
        if registros.count()>0:
            self.status = 1
            mensaje = "Encontrado"
            for u in registros:
                output.append({'hora': u['Hora'], 'dia': u['Dia'], 'fecha_adicional': u['Fecha_Adicional'],'codigo': u['Codigo'],'practica': u['practica'],'asistencia': u['Asistencia'], 'estadoElemento': u['Estado_Elemento'],'placaElemento': u['Placa_Elemento'], 'observacionesGenerales': u['Observaciones_Generales'],'fecha_reserva': u['Fecha_reserva'], 'sala': u['Sala'], 'banco': u['Banco'], 'elemento': u['Elemento'],'usuario': u['Usuario'], 'estado': u['Estado'], "monitor": u['Monitor']})
        else:
            self.status = 2
            mensaje = "No hay adicionales registrados para el día " + str(fecha) + " en el laboratorio de " + str(sala) + " durante la franja de " + str(hora) + " - " + str(hora+2)
            
        return {'status': self.status, 'mensaje': mensaje, 'data': output}

    def buscarreservalabo(self):
        user = mongo.db.Prestamo
        output = []
        self.status = 1
        mensaje = "Encontrado"
        for u in user.find():
            output.append({'hora': u['Hora'], 'dia': u['Dia'], 'fecha_reserva': u['Fecha_reserva'], 'practica': u['practica'], 'fecha_adicional': u['Fecha_Adicional'],
                           'sala': u['Sala'], 'banco': u['Banco'], 'elemento': u['Elemento'], 'codigo': u['Codigo'], 'usuario': u['Usuario'], 'estado': u['Estado']})
        return {'status': self.status, 'mensaje': mensaje, 'data': output}

    def borrarreserva(self):
        user = mongo.db.Prestamo
        codigo = request.json['codigo']
        sala = request.json['sala']
        banco = request.json['banco']
        fecha_adicional = request.json['fecha_adicional']
        hora = request.json['hora']
        mensaje = "Operación fallida"
        print(codigo)
        print(sala)
        print(banco)
        print(fecha_adicional)
        print(hora)
        borrar = user.delete_one({"$and": [{'Codigo': codigo}, {'Sala': sala}, {'Banco': banco}, {'Fecha_Adicional':fecha_adicional}, {'Hora':hora}]})
        if borrar:
            print("Entro al if")
            self.status = 1
            print("Pasó el Status:",self.status)
            mensaje = "usuarios eliminados"
        print("el mensaje: ",mensaje)
        return mensaje

    def editarreserva(self):
        user = mongo.db.Prestamo
        codigo = request.json['codigo']
        dia = request.json['dia']
        hora = request.json['hora']
        sala = request.json['sala']
        fecha_adicional = request.json['fecha_adicional']
        fecha_reserva = request.json['fecha_reserva']
        banco = request.json['banco']   
        elemento = request.json['elemento']     
        estadoElemento = request.json['estadoElemento']     
        placaElemento = request.json['placaElemento']     
        observacionesGenerales = request.json['observacionesGenerales']     

        mensaje = "Reserva modificada"
        actualizardatos = user.update({"$and": [{'Sala': sala}, {'Hora': hora}, {'Dia': dia}, {'Fecha_Adicional': fecha_adicional}, {
                                      'Fecha_reserva': fecha_reserva}, {'Banco': banco}, {'Codigo': codigo}]}, {"$set": {'Elemento': elemento,'Estado_Elemento':estadoElemento,'Placa_Elemento':placaElemento,'Observaciones_Generales':observacionesGenerales}})
        print(actualizardatos)
        if actualizardatos:
            self.status = 1
            mensaje
        return mensaje
    
    def editarAdicional(self):
        user = mongo.db.Prestamo
        datos = request.json['datos']
        for item in datos:
            user.update({"$and": [{'Codigo': item['codigo']}, {'Hora': item['hora']}, {'Dia': item['dia']}, 
                                  {'Fecha_Adicional': item['fecha_adicional']}, {'Fecha_reserva': item['fecha_reserva']}, 
                                  {'Banco': item['banco']}, {'Codigo': item['codigo']}]}, 
                        {"$set": {'Elemento': item['elemento'], 'Estado_Elemento':item['estadoElemento'], 
                                  'Placa_Elemento': item['placaElemento'], 'Asistencia': item['asistencia'],
                                  'Monitor': item['monitor'], 'Observaciones_Generales': item['observacionesGenerales']}})

        self.status = 1
        mensaje = "Reserva modificada"
        return mensaje

    def aprobarreserva(self):
        user = mongo.db.Prestamo
        hora = request.json['hora']
        sala = request.json['sala']
        fecha_adicional = request.json['fecha_adicional']
        banco = request.json['banco']        
        mensaje = "Operación realizada"
        aprobar = request.json['aprobar']
        print("APROBAR: ",type(aprobar))
        aprobado = "APROBADO"
        cancelado = "CANCELADO"
        print("aprobado: ",type(aprobado))
        print(hora,sala,fecha_adicional,banco,aprobar,mensaje)
        actualizardatos = user.update({"$and": [{'Sala': sala}, {'Hora': hora}, {'Fecha_Adicional': fecha_adicional},
                                      {'Banco': banco}]}, {"$set": {"Estado": aprobar}})
        print(actualizardatos)
        if actualizardatos:
            self.status = 1
            if (aprobar==aprobado):
                mensaje = "PRÁCTICA APROBADA"
            if (aprobar==cancelado):
                mensaje = "PRÁCTICA CANCELADA"

        return mensaje

    def asistenciareserva(self):
        user = mongo.db.Prestamo
        hora = request.json['hora']
        sala = request.json['sala']
        fecha_adicional = request.json['fecha_adicional']
        banco = request.json['banco']        
        mensaje = "Operación realizada"
        asistencia = request.json['aprobar']
        print("ASISTIÓ: ",type(asistencia))

        aprobado = "SI"
        cancelado = "NO"
        print("aprobado: ",type(asistencia))
        print(hora,sala,fecha_adicional,banco,asistencia,mensaje)
        actualizardatos = user.update({"$and": [{'Sala': sala}, {'Hora': hora}, {'Fecha_Adicional': fecha_adicional},
                                      {'Banco': banco}]}, {"$set": {"Asistencia": asistencia}})
        print(actualizardatos)
        if actualizardatos:
            self.status = 1
            if (asistencia==aprobado):
                mensaje = "EL USUARIO -SI- ASISTIÓ"
            if (asistencia==cancelado):
                mensaje = "EL USUARIO -NO- ASISTIÓ"

        return mensaje

    def send_mail(self):            
        usuario = request.json['usuario']
        fecha = request.json['fecha_adicional']
        laboratorio = request.json['laboratorio']
        horas = request.json['hora']
        estado = request.json['opcion']
        print("Aquí el estado de la petición: ")
        print(estado)

        email = "adminsalas@udistrital.edu.co"    
        sender="rjchiaj@correo.udistrital.edu.co"

        subject = "Estado de la reserva de adicional."
        msg = "Se le informa al estudiante " + usuario + " que el estado de la reserva adicional para el día " + fecha + " en el laboratorio de " + laboratorio + " a las " + str(horas) + " horas " + "es: "+ estado+". Para mayor información consultar la página: https://labing.udistrital.edu.co/"
        message = Message(subject,recipients=[email],sender = ('Richar Chia', sender),extra_headers={'Disposition-Notification-To': sender})
        message.body = msg
        
        with app.open_resource('PruebaAdjuntos.txt') as archivo:
            message.attach('PruebaAdjuntos.txt', 'text/txt', archivo.read())

        mail.send(message)
        self.status = 1
        mensaje = "Email enviado correctamente"
        return mensaje


    def buscarhorarios(self):
        user = mongo.db.Horario
        output = []
        self.status = 1
        mensaje = "Encontrado"
        for u in user.find():
            output.append({'asignatura': u['Asignatura'], 'docente': u['Docente'], 'proyecto': u['Proyecto'], 'grupo': u['Grupo'],
                           'dia': u['Dia'], 'horario': u['Horario'], 'sala': u['Sala'], 'fecha': u['Fecha ']})
        return {'status': self.status, 'mensaje': mensaje, 'data': output}

    def consultartipo(self):
        print("CONSULTAR TIPO")
        mensaje = "error"
        usuario = mongo.db.Usuarios
        codigo = request.json['codigo']
        output = []    
        self.status = 1
        mensaje = "Encontrado"
        for u in usuario.find({"Código_usuario":codigo}):
            output.append({"tipo": u['Tipo']})
        print(output)
        return {'status': self.status, 'mensaje': mensaje, 'data': output}

    def consultarLabo(self):
        usuario = mongo.db.Usuarios
        tipo = "Laboratorista"
        output = []
        self.status = 1
        mensaje = "Encontrado"
        for u in usuario.find({"Tipo": tipo}):
            output.append({'usuario': u['Nombre_usuario'], 'correo': u['Correo_usuario'],
                           'estado': u['Código_Estado'], 'identificacion': u['Código_usuario']})
        return {'status': self.status, 'mensaje': mensaje, 'data': output}

    def consultaeditlabo(self):
        print("HOLA AUXILIO")#BORRAR ESTO
        usuario = mongo.db.Usuarios
        codigo = request.json['codigo'] #NO ESTA LLEGANDO EL CÓDIGO DEL USUARIO        
        tipo = request.json['tipo']
        output = []
        self.status = 1
        mensaje = "Encontrado"
        print("HOLA AUXILIO 222")#BORRAR ESTO
        print([request.json['codigo'],codigo,tipo])#BORRAR ESTO


        for u in usuario.find({"$and": [{"Código_usuario": codigo}, {"Tipo": tipo}]}):
            print("11111POR")#BORRAR ESTO
            output.append(
                {'usuario': u['Nombre_usuario'], 'correo': u['Correo_usuario']})



        print("HOLA AUXILIO 333 CAMBIO")#BORRAR ESTO
        return {'status': self.status, 'mensaje': mensaje, 'data': output}

    def consultaeditadmin(self):
        usuario = mongo.db.Usuarios
        codigo = request.json['codigo']
        tipo = "Administrador"
        output = []
        self.status = 1
        mensaje = "Encontrado"
        for u in usuario.find({"$and": [{"Código_usuario": codigo}, {"Tipo": tipo}]}):
            output.append(
                {'usuario': u['Nombre_usuario'], 'correo': u['Correo_usuario']})
        return {'status': self.status, 'mensaje': mensaje, 'data': output}

    def consultarEquipo(self):
        print("ENTRA A CONSULTAR EQUIPO")
        elemento = mongo.db.Elementos
        sede = "FICC01"
        output = []
        self.status = 1
        encontrar = elemento.find({"Id_Sede": sede})
        if encontrar:
            mensaje = "Equipo Encontrado"
            for u in elemento.find({"Id_Sede": sede}):
                output.append({'placa': u['Placa'], 'nombreEquipo': u['Nombre_Del_Equipo'], 'estado': u['Estado'],'serie':u['Serie'], 'modelo':u['Referencia/Modelo'],
                               'marca': u['Marca'],  'numeroInterno': u['Codigo_Interno'], 'sede':u['Sede'], 'idSede':u['Id_Sede'],'dependencia':u['Dependencia'],'idUbicacion':u['Id_Ubicacion'],
                               'espacioFisico': u['Espacio_Fisico'], 'proveedor':u['Proveedor'], 'valorElemento':u['Valor_Elemento'], 'nit':u['Nit'],'ivaAplicado':u['Iva_Aplicado'],
                               'tipoContrato':u['Tipo_De_Contrato'],'totalValorelemento':u['Total_Valor_Elemento'], 'cantidadAsignada':u['Cantidad_Asignada'],'vigencia':u['Vigencia'],
                               'fechaAdquisicion':u['Fecha_De_Adquisicion'], 'numeroContrato':u['Numero_De_Contrato'], 'numeroFactura':u['#_De_Factura_Compra'], 'tiempoGarantia':u['Tiempo_De_Garantia'],
                               'frecuenciaMantenimiento':u['Frecuencia_De_Mantenimiento'],'manual':u['Cuenta_Con_Manual'],'tiempoVidautil':u['Tiempo_De_Vida_Util'],'tipoUso':u['Tipo_De_Uso'],
                               'potenciaElectrica':u['Potencia_Electrica'], 'tipoUso_otro':u['Tipo_De_Uso_Otro'], 'paisOrigen':u['Pais_De_Origen'],'impactoEquipo':u['Impacto_Uso_Del_Equipo'],
                               'especificacionesTecnicas':u['Especificaciones_Tecnicas'], 'accesorios' :u['Accesorios'], 'documentoFuncionario':u['Documento_Funcionario'], 'nombreFuncionario':u['Nombre_Del_Funcionario']
                               })
        else:
            mensaje = "no se encontro"
        return {'status': self.status, 'mensaje': mensaje, 'data': output}

    def editarequipo(self):
        elemento = mongo.db.Elementos
        placa = request.json['placa']
        sala = request.json['sala']
        iUbicacion = request.json['iUbicacion']
        mensaje = "Ubicación del equipo actualizada"
        actualizardatos = elemento.update({"Placa": placa}, {"$set": {
            "Espacio_Fisico": sala, 'Id_Ubicacion': iUbicacion}})
        if actualizardatos:
            self.status = 1
        else:
            self.status = 0
        return {'status': self.status, 'mensaje': mensaje}

    def desactivarEquipo(self):
        elemento = mongo.db.Elementos
        placa = request.json['placa']
        estado = "NO ACTIVO"
        mensaje = "Desactivado"
        desactivar = elemento.update({"Placa": placa}, {
            "$set": {"Estado": estado}})
        if desactivar:
            self.status = 1
            mensaje
        return mensaje

    def activarEquipo(self):
        elemento = mongo.db.Elementos
        placa = request.json['placa']
        estado = "ACTIVO"
        mensaje = "Activado"
        activar = elemento.update({"Placa": placa}, {
            "$set": {"Estado": estado}})
        if activar:
            self.status = 1
            mensaje
        return mensaje

    def registrarEquipo(self):
        mensaje = "0"
        elemento = mongo.db.Elementos
        placa = request.json['placa']
        serial = request.json['serial']
        nombreEquipo = request.json['nombreEquipo']
        serie = request.json['serie']
        estado = request.json['estado']
        modelo = request.json['modelo']
        marca = request.json['marca']
        numeroInterno = request.json['numeroInterno']
        sede = request.json['sede']
        idSede = request.json['idSede']
        dependencia = request.json['dependencia']
        idUbicacion = request.json['idUbicacion']
        espacioFisico = request.json['espacioFisico']
        proveedor = request.json['proveedor']
        valorElemento = request.json['valorElemento']
        nit = request.json['nit']
        ivaAplicado = request.json['ivaAplicado']
        tipoContrato = request.json['tipoContrato']
        totalValorelemento = request.json['totalValorelemento']
        cantidadAsignada = request.json['cantidadAsignada']
        vigencia = request.json['vigencia']
        fechaAdquisicion = request.json['fechaAdquisicion']
        numeroContrato = request.json['numeroContrato']
        numeroFactura = request.json['numeroFactura']
        tiempoGarantia = request.json['tiempoGarantia']
        frecuenciaMantenimiento = request.json['frecuenciaMantenimiento']
        manual = request.json['manual']
        tiempoVidautil = request.json['tiempoVidautil']
        tipoUso = request.json['tipoUso']
        potenciaElectrica = request.json['potenciaElectrica']
        tipoUso_otro = request.json['tipoUso_otro']
        paisOrigen = request.json['paisOrigen']
        impactoEquipo = request.json['impactoEquipo']
        especificacionesTecnicas = request.json['especificacionesTecnicas']
        accesorios = request.json['accesorios']
        documentoFuncionario = request.json['documentoFuncionario']
        nombreFuncionario = request.json['nombreFuncionario']
        laboratorista = request.json['laboratorista']
        pregunta = elemento.find_one({"Placa": placa})
        if pregunta:
            mensaje = "Este equipo ya existe en el sistema"
            self.status = 1
        else:
            nuevoElemento = elemento.insert({'Placa': placa, 'Serial': serial, 'Nombre_Del_Equipo': nombreEquipo,
                                             'Serie': serie, 'Estado': estado,
                                             'Referencia/Modelo': modelo, 'Marca': marca, 'Codigo_Interno': numeroInterno, 
                                             'Sede': sede, 'Id_Sede': idSede, 'Dependencia': dependencia, 'Id_Ubicacion': idUbicacion,
                                             'Espacio_Fisico': espacioFisico, 'Proveedor': proveedor, 'Valor_Elemento': valorElemento,
                                             'Nit': nit, 'Iva_Aplicado': ivaAplicado, 'Tipo_De_Contrato': tipoContrato, 'Total_Valor_Elemento': totalValorelemento,
                                             'Cantidad_Asignada': cantidadAsignada,
                                             'Vigencia': vigencia, 'Fecha_De_Adquisicion': fechaAdquisicion, 'Numero_De_Contrato': numeroContrato,
                                             '#_De_Factura_Compra': numeroFactura, 'Tiempo_De_Garantia': tiempoGarantia, 'Frecuencia_De_Mantenimiento': frecuenciaMantenimiento,
                                             'Cuenta_Con_Manual': manual, 'Tiempo_De_Vida_Util': tiempoVidautil, 'Tipo_De_Uso': tipoUso, 'Potencia_Electrica': potenciaElectrica,
                                             'Tipo_De_Uso_Otro': tipoUso_otro, 'Pais_De_Origen': paisOrigen, 'Impacto_Uso_Del_Equipo': impactoEquipo,
                                             'Especificaciones_Tecnicas': especificacionesTecnicas, 'Accesorios': accesorios,'Documento_Funcionario': documentoFuncionario,
                                             'Nombre_Del_Funcionario': nombreFuncionario, 'Laboratorista':laboratorista})
            mensaje = "Equipo Registrado correctamente"
            self.status = 1
        return mensaje

    def registrarMant(self):
        mensaje = "0"
        elemento = mongo.db.Mantenimiento
        placa = request.json['placa']
        Numero_Interno = request.json['Numero_Interno']
        fecha = request.json['fecha']
        codUsu = request.json['codUsu']
        nomUsu = request.json['nomUsu']
        sala = request.json['sala']
        Horario = request.json['Horario']
        descDano = request.json['descDano']
        Tipo_De_Mantenimiento = request.json['Tipo_De_Mantenimiento']
        fechaReal = request.json['fechaReal']
        nomEmpresa = request.json['nomEmpresa']
        nit = request.json['nit']		
        tiempoGar = request.json['tiempoGar']		
        espMantReal = request.json['espMantReal']        
        costMant = request.json['costMan']
        numOrdServ = request.json['numOrdServ']                                 
        vigencia = request.json['vigencia']
        supervisor = request.json['supervisor']
        repuestos = request.json['repuestos']
        proxMant = request.json['proxMant'] 
        observaciones = request.json['observaciones']
        nomEqu = request.json['nomEqu']        
        Estado = request.json['Estado']
        print(nomEmpresa,nit,tiempoGar,espMantReal,supervisor,fechaReal)
        nuevoElemento = elemento.insert({'Placa':placa,'Numero_Interno':Numero_Interno,'Fecha':fecha,'Codigo_Usuario ':codUsu,'Nombre_Usuario':nomUsu,'Sala':sala,'Horario':Horario,'Descripcion_Dano':descDano,'Tipo_De_Mantenimiento':Tipo_De_Mantenimiento,'Fecha_De_Realizacion':fechaReal,'Nombre_Empresa_Contratada':nomEmpresa,'Nit':nit,'Tiempo_De_Garantia':tiempoGar,'Especificaciones_Del_Mantenimiento_Realizado':espMantReal,'Costo_Mantenimiento':costMant,'Numero_Orden_De_Servicio_Del_Mantenimiento':numOrdServ,'Vigencia':vigencia,'Supervisor':supervisor,'Repuestos':repuestos,'Proximo_Mantenimiento':proxMant,'Observaciones':observaciones,'Estado':Estado})
        mensaje = "Mantenimiento Registrado correctamente"
        self.status = 1
        return mensaje


    def consultaMantenimiento(self):
        
        mantenimiento = mongo.db.Mantenimiento
        output = []
        self.status = 1
        encontrar = mantenimiento.find({})
        if encontrar:
            mensaje = "Mantenimientos Encontrados"
            for u in mantenimiento.find({}):
                output.append({'placa': u['Placa'], 'laboratorista': u['Nombre_Usuario'],'sala':u['Sala'], 'dano':u['Descripcion_Dano'], 'fechaMantenimiento':u['Fecha'],
                               'tipoMantenimiento': u['Tipo_De_Mantenimiento'],  'numeroInterno': u['Numero_Interno'], 'hora':u['Horario'], 'estado':u['Estado'],
                                'fechaReal':u['Fecha_De_Realizacion'], 'nomEmpresa':u['Nombre_Empresa_Contratada'],'NIT':u['Nit'],'tiempGar':u['Tiempo_De_Garantia'],'costMant':u['Costo_Mantenimiento'],
                                'numOrdServ':u['Numero_Orden_De_Servicio_Del_Mantenimiento'],'vigencia':u['Vigencia'],'supervisor':u['Supervisor'],'repuestos':u['Repuestos'],'proxMant':['Proximo_Mantenimiento'],
                                'espMantReal':u['Especificaciones_Del_Mantenimiento_Realizado'],'observaciones':u['Observaciones'],
                               })
        else:
            mensaje = "no se encontro"
        # print("SALIO AJI MANTENIMIENTOS:", mensaje)
        # print("DATOS: ",output)
        return {'status': self.status, 'mensaje': mensaje, 'data': output}    


    def consultaTraslados(self):
        traslado = mongo.db.Traslados
        output = []
        self.status = 1
        codigoLab = request.json['codigoLab']
        estado = "PENDIENTE"
        encontrar = traslado.find({})
        if encontrar:

            mensaje = "Traslados encontrados"
            for u in traslado.find({"$and": [{"Codigo_Laboratorista": codigoLab}, {"Estado": estado}]}):
                output.append({'placa': u['Placa'], 'nombreEquipo': u['Nombre_Equipo'],'salaOrigen': u['Sala_Origen'], 'salaDestino': u['Sala_Destino'], 'codigoLaboratorista': u['Codigo_Laboratorista'],
                               'laboratorista': u['Nombre_Laboratorista'],'estado':u['Estado'], 'ubicacionExacta' :u['Ubicacion_Exacta']
                               })
        else:
            mensaje = "No hay traslados pendientes"
            
        return {'status': self.status, 'mensaje': mensaje, 'data': output} 
            

    def buscarUno(self):
        mensaje = "0"
        user = mongo.db.Usuarios
        identificacion = request.json['identificacion']
        pregunta = user.find_one({"identificacion": identificacion})
        if pregunta:
            mensaje = "Encontrado"
            self.status = 1
            self.datos = {
                "usuario": pregunta['usuario'], "correo": pregunta['correo'], "telefono": pregunta['telefono']}
        else:
            mensaje = "No fue encontrado"
            self.status = 2
        return mensaje

    def borrarPrueba(self):
        usuario = mongo.db.Usuarios
        usuarioBusc = request.json['usuario']
        print(usuarioBusc)
        borrar = usuario.delete_one({'usuario': usuarioBusc})
        if borrar:
            self.status = 1
            mensaje = " usuarios eliminados"
        return mensaje

    def altiumLogo(self):
        return send_file(
                'Licencias/altium-logo.jpg',
                as_attachment=True,
                attachment_filename='Licencias/altium-logo.jpg',
                mimetype='image/jpeg'
                )

    def psimLogo(self):
        return send_file(
                'Licencias/psim-logo.jpg',
                as_attachment=True,
                attachment_filename='Licencias/psim-logo.jpg',
                mimetype='image/jpeg'
                )

    def solidWorksLogo(self):
        return send_file(
                'Licencias/solidworks-logo.jpg',
                as_attachment=True,
                attachment_filename='Licencias/solidworks-logo.jpg',
                mimetype='image/jpeg'
                )

    def xirioLogo(self):
        return send_file(
                'Licencias/xirio-logo.jpg',
                as_attachment=True,
                attachment_filename='Licencias/xirio-logo.jpg',
                mimetype='image/jpeg'
                )
    
    def licenciasEstudiantes(self):
        # Configuracion de las credenciales
        scope =["https://spreadsheets.google.com/feeds",
                'https://www.googleapis.com/auth/spreadsheets',
                "https://www.googleapis.com/auth/drive.file",
                "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name('Licencias/Sheets API key.json', scope)
        client = gspread.authorize(creds)

        # Se instancia la hoja del documento compartido
        worksheet = client.open("Prueba Licencias").worksheet("Hoja 1")

        form = request.json['form']

        date  = datetime.datetime.now()
        fecha = date.strftime("%d/%m/%Y")
        hora = date.strftime("%H:%M")
        
        data = [fecha, hora, form['nombre'], form['codigo'], form['correo'], form['documento'], 
                form['proyecto'], form['facultad'], form['software'], "", form['tipoUsuario'],
                form['motivo'], "", "", ""]

        worksheet.insert_row(data,2)
        mensaje = "Licencia agregada "
        return mensaje
    
    def addHorarioAdicional(self):
        laboratorio = request.json['name']
        bancos = request.json['bancos']
        dia = request.json['dia']
        hora = request.json['hora']
        horario = mongo.db.horariosAdicionales
        pregunta = horario.find_one({'Laboratorio': laboratorio, 'Dia': dia, 'Hora': hora})
        if(pregunta):
            horario.update({"$and": [{'Laboratorio': laboratorio, 'Dia': dia, 'Hora': hora}]}, {"$set": {"Bancos": bancos}})
            mensaje = "Adicional modificado"
        else:
            horario.insert({'Laboratorio': laboratorio, 'Bancos': bancos, 'Dia': dia, 'Hora': hora})
            mensaje = "Adicional agregado"
        return mensaje

    def deleteHorarioAdicional(self):
        laboratorio = request.json['name']
        bancos = request.json['bancos']
        dia = request.json['dia']
        hora = request.json['hora']
        horario = mongo.db.horariosAdicionales
        horario.delete_one({"$and": [{'Laboratorio': laboratorio}, {'Bancos': bancos}, {'Dia':dia}, {'Hora':hora}]})
        mensaje = "Adicional eliminado"
        return mensaje

    
    def getHorariosAdicionales(self):
        year = datetime.datetime.now().year
        week = datetime.datetime.now().isocalendar()[1]
        date_string = str(year) + "-W" + str(week)
        horario = mongo.db.horariosAdicionales
        output = []
        for item in horario.find():
            date = datetime.datetime.strptime(date_string + '-' + str(item['Dia']), '%Y-W%W-%w')
            start = date.strftime('%Y-%m-%d')+' '+str(item['Hora'])+":00"
            end = date.strftime('%Y-%m-%d')+' '+str(item['Hora']+2)+":00"
            output.append({'name': item['Laboratorio'], 'weekday': item['Dia'], 'bancos': item['Bancos'], 'start': start, 'end': end, 'hour': item['Hora']})        
        mensaje = "Exito"
        self.status = 1
        
        return {'status': self.status, 'mensaje': mensaje, 'data': output}    

    ############### Metodos para crear documentos Excel ###########################
    def excelTemporalElementos(self):
        # Este if pregunta si existe la base de datos temporal y en caso de que si, la elimina
        if 'temporalExcel' in mongo.db.list_collection_names():
            temporalExcel = mongo.db.temporalExcel
            temporalExcel.drop()

        # Valida los campos 'true' enviados por el usuario para crear una base de datos temporal. Si el campo itemPost es verdadero, recorre todos la coleccion elementos y solo agrega el campo solicitado a la coleccion temporal. El if mas interno es para validar si no existe un documento en la coleccion con ese id para insertar un nuevo elemento o en caso contrario actualizar uno
        temporalExcel = mongo.db["temporalExcel"]
        elemento = mongo.db.Elementos
        req_data = request.get_json()
        for itemPost in req_data:
            if request.json[itemPost]:
                for (index,item) in enumerate(elemento.find().limit(20)):
                    if temporalExcel.count_documents({"_id": index+1})==0:
                        temporalExcel.insert_one({"_id":index+1,itemPost:item[itemPost]})
                    else:
                        temporalExcel.update({"_id":index+1},{'$set':{itemPost:item[itemPost]}})

        self.status = 1    
        mensaje = "Se creo la base de datos temporal"    
        return mensaje
    
    def excelTemporalFilterDate(self):
        # Se toman las fechas y se separan los caracteres en un vector [dia, mes, año]
        startDate = request.json['startDate'].split('/')
        endDate = request.json['endDate'].split('/')

        # Variable que tiene asociada el nombre del elemento asociado a la fecha en la base de datos
        nameKeyDate = request.json['nameKeyDate']

        # Se borra la base de datos temporal
        if 'temporalExcelFilterDate' in mongo.db.list_collection_names():
            temporalExcelFilterDate = mongo.db.temporalExcelFilterDate
            temporalExcelFilterDate.drop()

        # Se crea la base de datos temporal
        temporalExcelFilterDate = mongo.db["temporalExcelFilterDate"]

        # Se obtiene la base de datos segun el nombre que llega en la peticion
        elemento = mongo.db.get_collection(request.json['dataBase'])

        for (index,item) in enumerate(elemento.find().limit(35)):
            # La fecha tiene un espacio...
            # Se toma la fecha del elemento y se separa en un vector [dia,mes,año]
            dateElement = item[nameKeyDate].split('/')
            # Se valida si la fecha del elemento esta en el rango de busqueda
            if startDate[2] <= dateElement[2] and dateElement[2] <= endDate[2] and startDate[1] <= dateElement[1] and dateElement[1] <= endDate[1] and startDate[0] <= dateElement[0] and dateElement[0] <= endDate[0]:
                temporalExcelFilterDate.insert_one(item)

        if (temporalExcelFilterDate.find().count()!=0):
            self.status = 1    
            mensaje = "Se creo la base de datos temporal"    
        else:
            self.status = 3   
            mensaje = "No hay elementos que coincidan con los datos de busqueda."    

        return mensaje
    
    def excelTemporalFilterDateBancosSala(self):
        # Se toman las fechas y se separan los caracteres en un vector [dia, mes, año]
        startDate = request.json['startDate'].split('/')
        endDate = request.json['endDate'].split('/')

        # Variable que tiene asociada el nombre del elemento asociado a la fecha en la base de datos
        nameKeyDate = request.json['nameKeyDate']

        # Se borra la base de datos temporal
        if 'temporalExcelFilterDate' in mongo.db.list_collection_names():
            temporalExcelFilterDate = mongo.db.temporalExcelFilterDate
            temporalExcelFilterDate.drop()

        # Se crea la base de datos temporal
        temporalExcelFilterDate = mongo.db["temporalExcelFilterDate"]

        # Se obtiene la base de datos segun el nombre que llega en la peticion
        elemento = mongo.db.get_collection(request.json['dataBase'])

        for (index,item) in enumerate(elemento.find().limit(35)):
            # La fecha tiene un espacio...
            # Se toma la fecha del elemento y se separa en un vector [dia,mes,año]
            dateElement = item[nameKeyDate].split('/')
            # Se valida si la fecha del elemento esta en el rango de busqueda
            if startDate[2] <= dateElement[2] and dateElement[2] <= endDate[2] and startDate[1] <= dateElement[1] and dateElement[1] <= endDate[1] and startDate[0] <= dateElement[0] and dateElement[0] <= endDate[0]:
                if item["Sala"] == request.json["sala"]:
                    temporalExcelFilterDate.insert_one(item)

        if (temporalExcelFilterDate.find().count()!=0):
            self.status = 1    
            mensaje = "Se creo la base de datos temporal"    
        else:
            self.status = 3   
            mensaje = "No hay elementos que coincidan con los datos de busqueda."    

        return mensaje
    
    def obtenerElementosExcel(self):
        elemento = mongo.db.Elementos
        sede = "FICC01"
        output = []
        self.status = 1
        encontrar = elemento.find({"Id_Sede": sede})
        if encontrar:
            mensaje = "Base de datos de elementos encontrada"
            for u in elemento.find({"Id_Sede": sede}):
                output.append({'Placa': u['Placa'], 'Nombre_Del_Equipo': u['Nombre_Del_Equipo'], 'Estado': u['Estado'],'Serie':u['Serie'], 'Referencia/Modelo':u['Referencia/Modelo'],                            'Marca': u['Marca'],  'Codigo_Interno': u['Codigo_Interno'], 'Sede':u['Sede'], 'Id_Sede':u['Id_Sede'],'Dependencia':u['Dependencia'],'Id_Ubicacion':u['Id_Ubicacion'],
                               'Espacio_Fisico': u['Espacio_Fisico'], 'Proveedor':u['Proveedor'], 'Valor_Elemento':u['Valor_Elemento'], 'Nit':u['Nit'],'Iva_Aplicado':u['Iva_Aplicado'],
                               'Tipo_De_Contrato':u['Tipo_De_Contrato'],'Total_Valor_Elemento':u['Total_Valor_Elemento'], 'Cantidad_Asignada':u['Cantidad_Asignada'],'Vigencia':u['Vigencia'],
                               'Fecha_De_Adquisicion':u['Fecha_De_Adquisicion'], 'Numero_De_Contrato':u['Numero_De_Contrato'], '#_De_Factura_Compra':u['#_De_Factura_Compra'], 'Tiempo_De_Garantia':u['Tiempo_De_Garantia'],
                               'Frecuencia_De_Mantenimiento':u['Frecuencia_De_Mantenimiento'],'Cuenta_Con_Manual':u['Cuenta_Con_Manual'],'Tiempo_De_Vida_Util':u['Tiempo_De_Vida_Util'],'Tipo_De_Uso':u['Tipo_De_Uso'],
                               'Potencia_Electrica':u['Potencia_Electrica'], 'Tipo_De_Uso_Otro':u['Tipo_De_Uso_Otro'], 'Pais_De_Origen':u['Pais_De_Origen'],'Impacto_Uso_Del_Equipo':u['Impacto_Uso_Del_Equipo'],
                               'Especificaciones_Tecnicas':u['Especificaciones_Tecnicas'], 'Accesorios' :u['Accesorios'], 'Documento_Funcionario':u['Documento_Funcionario'], 'Nombre_Del_Funcionario':u['Nombre_Del_Funcionario']
                               })
        else:
            mensaje = "no se encontro"
        return {'status': self.status, 'mensaje': mensaje, 'data': output}

    def createExcel(self):
        # Se crea la instancia para crear el excel
        xlsx = Workbook()
        # Se activa la primera hoja del documento
        doc = xlsx.active
        # Por defecto la primera hoja se llama Sheet, entonces se le cambia a Informe
        doc.title = "Informe"

        header = []
        data = []

        elemento = mongo.db.temporalExcel

        for item in elemento.find().limit(1):
            for nameKey in item:
                if nameKey != "_id":
                    header.append(nameKey)

        for item in elemento.find():
            dataAux = []
            for nameKey in item:
                if nameKey != "_id":
                    dataAux.append(item[nameKey])
            data.append(dataAux)
        
        if 'temporalExcel' in mongo.db.list_collection_names():
            temporalExcel = mongo.db.temporalExcel
            temporalExcel.drop()

        # Se agrega el encabezado y el contenido de la tabla. Se agrega fila por fila
        doc.append(header)    
        for row in data:
            doc.append(row)

        # ESTILOS Header
        for col in doc.iter_cols(min_col=1,max_col=doc.max_column,min_row=1,max_row=1):
            for col_index in col:
                col_index.alignment = Alignment(horizontal='center',vertical='center') 

        # Para desplazar la tabla se insertan nuevas filas y columnas en blanco
        marginLeft = 1
        marginTop = 1
        doc.insert_cols(1,marginLeft)
        doc.insert_rows(1,marginTop)

        # PosX es la celda en que la tabla inicia (esquina superior izquierda), esto es: la cantidad de celdas del margen izquierdo : la cantidad de celdas del margen superior
        # PosY es la celda en que la tabla termina (esquina inferior derecha), esto es: la ultima columna con datos : la ultima fila con datos
        posX = get_column_letter(marginLeft+1)+str(marginTop+1)
        posY = get_column_letter(doc.max_column)+str(doc.max_row)

        # Se le da el formato de tabla para establecer los estilos
        table = Table(displayName="Tabla1", ref=posX+":"+posY)

        style = TableStyleInfo(name="TableStyleLight10", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True)

        # Se aplican los estilos a la tabla creada
        table.tableStyleInfo = style

        # Se añade la tabla al documento
        doc.add_table(table)

        # Se juntan las celdas para crear el header
        # doc.merge_cells('B2:I2')
        # doc.merge_cells('B3:I3')
        # doc.merge_cells('B4:I4')

        # Se inserta el header
        # doc["B2"]="MANTENIMIENTOS LABORATORIOS DE INGENIERIA"
        # doc["B3"]="SALAS DE INFORMATICA TRIMESTRE II"
        # doc["B4"]="PERIODO COMPRENDIDO ENTRE EL 1 DE ABRIL HASTA EL 30 DE ABRIL DEL 2020"
            
        return Response(
            save_virtual_workbook(xlsx),
            headers={
            'Content-Disposition': 'attachment; filename=Prueba.xlsx',
            'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

    def excelHorarios_DiasSemanaVsHoras(self):
        elemento = mongo.db.temporalExcelFilterDate
        
        xlsx = Workbook()
        doc = xlsx.active
        doc.title = "InformeHorarios"    

        header = ["Asignatura","Docente","Grupo","Sala","Dia","Horario","Fecha"]
        data = []

        for item in elemento.find().limit(35):
            data.append([item["Asignatura"],item["Docente"],item["Grupo"],item["Sala"],item["Dia"],item["Horario"],item["Fecha "]])
        
        doc.append(header)    
        for row in data:
            doc.append(row)

        for col in doc.iter_cols(min_col=1,max_col=doc.max_column,min_row=1,max_row=1):
            for col_index in col:
                col_index.alignment = Alignment(horizontal='center',vertical='center') 

        marginLeft = 1
        marginTop = 1
        doc.insert_cols(1,marginLeft)
        doc.insert_rows(1,marginTop)

        posX = get_column_letter(marginLeft+1)+str(marginTop+1)
        posY = get_column_letter(doc.max_column)+str(doc.max_row)

        table = Table(displayName="Tabla1", ref=posX+":"+posY)

        style = TableStyleInfo(name="TableStyleLight10", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True)

        table.tableStyleInfo = style

        doc.add_table(table)

        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

        inicioTablaGrafico = doc.max_column+4
        finalTablaGrafico = inicioTablaGrafico+4

        posInicial = get_column_letter(inicioTablaGrafico) + str(2)
        posFinal = get_column_letter(finalTablaGrafico) + str(2)

        doc[posInicial].border = thin_border
        doc.merge_cells(posInicial+":"+posFinal)
        doc[posInicial] = "HORAS DE CLASE POR DIA DE LA SEMANA"
        doc[posInicial].alignment = Alignment(horizontal='center',vertical='center')
        
        dias = ["Lunes","Martes","Miercoles","Jueves","Viernes"]
        cantidad_dias = []
        
        for row in doc.iter_rows(min_col=inicioTablaGrafico,max_col=finalTablaGrafico,min_row=3,max_row=3):
            for (index,cell) in enumerate(row):
                cell.value = dias[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border
                cantidad_dias.append(elemento.find({"Dia":dias[index]}).count()*2)

        for row in doc.iter_rows(min_col=inicioTablaGrafico,max_col=finalTablaGrafico,min_row=4,max_row=4):
            for (index,cell) in enumerate(row):
                cell.value = cantidad_dias[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border

        labels = Reference(doc,min_col=inicioTablaGrafico,min_row=3,max_col=finalTablaGrafico)      
        data = Reference(doc,min_col=inicioTablaGrafico,min_row=4,max_col=finalTablaGrafico)      

        chart = BarChart()
        chart.type = "bar" 
        chart.y_axis.title = "Cantidad de horas"  
        chart.title = "Cantidad de horas al día por la semana"
        chart.add_data(data,from_rows=True)
        chart.set_categories(labels)
        chart.legend = None
        chart.varyColors = True
        doc.add_chart(chart,"J6")  

        tab2 = xlsx.create_sheet("SalasVsHoras")    
        salas = ["500","501","502","503","504"]
        cantidad_horas = []

        tab2['B2'].border = thin_border
        tab2['B2'].alignment = Alignment(horizontal='center',vertical='center')
        tab2['B2'] = "HORAS DE CLASE A LA SEMANA POR SALA"
        tab2.merge_cells('B2:F2')

        for row in tab2.iter_rows(min_col=2,max_col=6,min_row=3,max_row=3):
            for (index,cell) in enumerate(row):
                cell.value = salas[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border
                cantidad_horas.append(elemento.find({"Sala":salas[index]}).count()*2)

        for row in tab2.iter_rows(min_col=2,max_col=6,min_row=4,max_row=4):
            for (index,cell) in enumerate(row):
                cell.value = cantidad_horas[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border

        labels = Reference(tab2,min_col=2,max_col=6,min_row=3)
        data = Reference(tab2,min_col=2,max_col=6,min_row=4)      

        chart = BarChart()
        chart.type = "col"
        chart.title = "Cantidad de horas a la semana por sala"
        chart.add_data(data,from_rows=True)
        chart.set_categories(labels)
        chart.legend = None
        chart.varyColors = True
        chart.x_axis.title = "Número de la sala"  
        chart.y_axis.title = "Cantidad de horas"  
        # chart.style = 7
        tab2.add_chart(chart,"B6")         

        if 'temporalExcelFilterDate' in mongo.db.list_collection_names():
            temporalExcelFilterDate = mongo.db.temporalExcelFilterDate
            temporalExcelFilterDate.drop()

        return Response(
            save_virtual_workbook(xlsx),
            headers={
            'Content-Disposition': 'attachment; filename=Prueba.xlsx',
            'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
    
    def excelPrestamos(self):
        elemento = mongo.db.temporalExcelFilterDate
        
        xlsx = Workbook()
        tab1 = xlsx.active
        tab1.title = "InformePrestamos"    

        header = ["Usuario","Código","Sala","Banco","Día","Hora","Fecha adicional","Fecha de reserva","Estado"]
        data = []

        for item in elemento.find():
            data.append([item["Usuario"],item["Codigo"],item["Sala"],item["Banco"],item["Dia"],item["Hora"],item["Fecha_Adicional"],item["Fecha_reserva"],item["Estado"]])
        
        tab1.append(header) 
           
        for row in data:
            tab1.append(row)

        for col in tab1.iter_cols(min_col=1,max_col=tab1.max_column,min_row=1,max_row=1):
            for col_index in col:
                col_index.alignment = Alignment(horizontal='center',vertical='center') 

        marginLeft = 1
        marginTop = 1
        tab1.insert_cols(1,marginLeft)
        tab1.insert_rows(1,marginTop)

        posX = get_column_letter(marginLeft+1)+str(marginTop+1)
        posY = get_column_letter(tab1.max_column)+str(tab1.max_row)

        table = Table(displayName="Tabla1", ref=posX+":"+posY)

        style = TableStyleInfo(name="TableStyleLight10", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True)

        table.tableStyleInfo = style

        tab1.add_table(table)

        # Para ajustar estilos del borde de la celda en las tablas de los graficos
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


        # Segunda hoja, primer grafico: dias de la semana vs cantidad de prestamos
        tab2 = xlsx.create_sheet("Prestamos por día")    
        dias = ["Monday","Tuesday","Wednesday","Thursday","Friday"]
        cantidad_prestamos = []

        lengDatas = len(dias)+1

        tab2['B2'].border = thin_border
        tab2['B2'].alignment = Alignment(horizontal='center',vertical='center')
        tab2['B2'] = "Cantidad de prestamos por día de la semana"
        tab2.merge_cells('B2:'+get_column_letter(lengDatas)+'2')

        for row in tab2.iter_rows(min_col=2,max_col=lengDatas,min_row=3,max_row=3):
            for (index,cell) in enumerate(row):
                cell.value = dias[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border
                cantidad_prestamos.append(elemento.find({"Dia":dias[index]}).count())

        for row in tab2.iter_rows(min_col=2,max_col=lengDatas,min_row=4,max_row=4):
            for (index,cell) in enumerate(row):
                cell.value = cantidad_prestamos[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border

        labels = Reference(tab2,min_col=2,max_col=lengDatas,min_row=3)
        data = Reference(tab2,min_col=2,max_col=lengDatas,min_row=4)      

        chart = BarChart()
        chart.type = "col"
        chart.title = "Cantidad de prestamos por día de la semana"
        chart.add_data(data,from_rows=True)
        chart.set_categories(labels)
        chart.legend = None
        chart.varyColors = True
        chart.x_axis.title = "Día de la semana"  
        chart.y_axis.title = "Cantidad de prestamos"  
        tab2.add_chart(chart,"B6")         

        # Tercera hoja, segundo grafico: sala de laboratorio vs cantidad de prestamos
        tab3 = xlsx.create_sheet("Prestamos por sala")    
        salasLab = ["Circuitos","Comunicaciones","Digitales","Electronica A","Electronica B","Fisica 509","Fisica 510","Maquinas","Electronica Básica"]
        cantidad_prestamos = []

        lengDatas = len(salasLab)+1

        tab3['B2'].border = thin_border
        tab3['B2'].alignment = Alignment(horizontal='center',vertical='center')
        tab3['B2'] = "Prestamos por sala de laboratorio"
        tab3.merge_cells('B2:'+get_column_letter(lengDatas)+'2')

        for row in tab3.iter_rows(min_col=2,max_col=lengDatas,min_row=3,max_row=3):
            for (index,cell) in enumerate(row):
                cell.value = salasLab[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border
                cantidad_prestamos.append(elemento.find({"Sala":salasLab[index]}).count())

        for row in tab3.iter_rows(min_col=2,max_col=lengDatas,min_row=4,max_row=4):
            for (index,cell) in enumerate(row):
                cell.value = cantidad_prestamos[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border

        labels = Reference(tab3,min_col=2,max_col=lengDatas,min_row=3)
        data = Reference(tab3,min_col=2,max_col=lengDatas,min_row=4)      

        chart = BarChart()
        chart.type = "col"
        chart.title = "Cantidad de prestamos por sala de laboratorio"
        chart.add_data(data,from_rows=True)
        chart.set_categories(labels)
        chart.legend = None
        chart.varyColors = True
        chart.y_axis.title = "Cantidad de prestamos"  
        chart.x_axis.title = "Sala del laboratorio"
        tab3.add_chart(chart,"B6")

        # Cuarta hoja, tercer grafico: hora de prestamo vs cantidad de prestamos
        tab4 = xlsx.create_sheet("Prestamos por hora")    
        horas = ["6:00","8:00","10:00","12:00","14:00","16:00","18:00","20:00","22:00"]
        cantidad_prestamos = []

        lengDatas = len(horas)+1

        tab4['B2'].border = thin_border
        tab4['B2'].alignment = Alignment(horizontal='center',vertical='center')
        tab4['B2'] = "Prestamos por hora"
        tab4.merge_cells('B2:'+get_column_letter(lengDatas)+'2')

        for row in tab4.iter_rows(min_col=2,max_col=lengDatas,min_row=3,max_row=3):
            for (index,cell) in enumerate(row):
                cell.value = horas[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border
                cantidad_prestamos.append(elemento.find({"Hora":horas[index]}).count())

        for row in tab4.iter_rows(min_col=2,max_col=lengDatas,min_row=4,max_row=4):
            for (index,cell) in enumerate(row):
                cell.value = cantidad_prestamos[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border

        labels = Reference(tab4,min_col=2,max_col=lengDatas,min_row=3)
        data = Reference(tab4,min_col=2,max_col=lengDatas,min_row=4)      

        chart = BarChart()
        chart.type = "col"
        chart.title = "Cantidad de prestamos según el horario del adicional"
        chart.add_data(data,from_rows=True)
        chart.set_categories(labels)
        chart.legend = None
        chart.varyColors = True
        chart.y_axis.title = "Cantidad de prestamos"  
        chart.x_axis.title = "Horario del adicional"
        tab4.add_chart(chart,"B6")

        # Quinta hoja, cuarto grafico: estado del prestamo vs cantidad de prestamos
        tab5 = xlsx.create_sheet("Estado de los prestamos")    
        estados = ["APROBADO","PENDIENTE","CANCELADO","RECHAZADO"]
        cantidad_prestamos = []

        lengDatas = len(estados)+1

        tab5['B2'].border = thin_border
        tab5['B2'].alignment = Alignment(horizontal='center',vertical='center')
        tab5['B2'] = "Estado de los prestamos"
        tab5.merge_cells('B2:'+get_column_letter(lengDatas)+'2')

        for row in tab5.iter_rows(min_col=2,max_col=lengDatas,min_row=3,max_row=3):
            for (index,cell) in enumerate(row):
                cell.value = estados[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border
                cantidad_prestamos.append(elemento.find({"Estado":estados[index]}).count())

        for row in tab5.iter_rows(min_col=2,max_col=lengDatas,min_row=4,max_row=4):
            for (index,cell) in enumerate(row):
                cell.value = cantidad_prestamos[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border

        for row in tab5.iter_rows(min_col=2,max_col=lengDatas,min_row=5,max_row=5):
            for (index,cell) in enumerate(row):
                cell.value = cantidad_prestamos[index]/(sum(cantidad_prestamos))*100
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border

        labels = Reference(tab5,min_col=2,max_col=lengDatas,min_row=3)
        data = Reference(tab5,min_col=2,max_col=lengDatas,min_row=5)      

        chart = PieChart()        
        chart.title = "Estado de los prestamos"
        chart.add_data(data,from_rows=True)
        chart.set_categories(labels)
        chart.varyColors = True
        chart.legend.position = 'b'
        tab5.add_chart(chart,"B7")

        # Sexta hoja, quinto grafico: estado del prestamo vs cantidad de prestamos
        tab6 = xlsx.create_sheet("Uso de bancos en Electronica B")    
        bancos = ["Banco 1","Banco 2","Banco 3","Banco 4","Banco 5","Banco 6"]
        cantidad_prestamos = []  

        lengDatas = len(bancos)+1

        tab6['B2'].border = thin_border
        tab6['B2'].alignment = Alignment(horizontal='center',vertical='center')
        tab6['B2'] = "Uso de los bancos en Electrónica B"
        tab6.merge_cells('B2:'+get_column_letter(lengDatas)+'2')

        for row in tab6.iter_rows(min_col=2,max_col=lengDatas,min_row=3,max_row=3):
            for (index,cell) in enumerate(row):
                cell.value = bancos[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border
                cantidad_prestamos.append(elemento.find({"Banco":str(index+1),"Sala":"Electronica B"}).count())

        for row in tab6.iter_rows(min_col=2,max_col=lengDatas,min_row=4,max_row=4):
            for (index,cell) in enumerate(row):
                cell.value = cantidad_prestamos[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border

        labels = Reference(tab6,min_col=2,max_col=lengDatas,min_row=3)
        data = Reference(tab6,min_col=2,max_col=lengDatas,min_row=4)      

        chart = BarChart()
        chart.type = "col"
        chart.title = "Cantidad de prestamos para los bancos de Electronica B"
        chart.add_data(data,from_rows=True)
        chart.set_categories(labels)
        chart.legend = None
        chart.varyColors = True
        chart.y_axis.title = "Cantidad de prestamos"  
        chart.x_axis.title = "Bancos de la sala de Electrónica B"
        tab6.add_chart(chart,"B6")

        if 'temporalExcelFilterDate' in mongo.db.list_collection_names():
            temporalExcelFilterDate = mongo.db.temporalExcelFilterDate
            temporalExcelFilterDate.drop()

        return Response(
            save_virtual_workbook(xlsx),
            headers={
            'Content-Disposition': 'attachment; filename=Prueba.xlsx',
            'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

    def excelPrestamosBancosSala(self):
        elemento = mongo.db.temporalExcelFilterDate
        
        xlsx = Workbook()
        tab1 = xlsx.active
        tab1.title = "InformePrestamos"    

        header = ["Usuario","Código","Sala","Banco","Día","Hora","Fecha adicional","Fecha de reserva","Estado"]
        data = []

        for item in elemento.find():
            data.append([item["Usuario"],item["Codigo"],item["Sala"],item["Banco"],item["Dia"],item["Hora"],item["Fecha_Adicional"],item["Fecha_reserva"],item["Estado"]])
        
        tab1.append(header) 
           
        for row in data:
            tab1.append(row)

        for col in tab1.iter_cols(min_col=1,max_col=tab1.max_column,min_row=1,max_row=1):
            for col_index in col:
                col_index.alignment = Alignment(horizontal='center',vertical='center') 

        marginLeft = 1
        marginTop = 1
        tab1.insert_cols(1,marginLeft)
        tab1.insert_rows(1,marginTop)

        posX = get_column_letter(marginLeft+1)+str(marginTop+1)
        posY = get_column_letter(tab1.max_column)+str(tab1.max_row)

        table = Table(displayName="Tabla1", ref=posX+":"+posY)

        style = TableStyleInfo(name="TableStyleLight10", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True)

        table.tableStyleInfo = style

        tab1.add_table(table)

        # Para ajustar estilos del borde de la celda en las tablas de los graficos
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


        # Sexta hoja, quinto grafico: 
        tab2 = xlsx.create_sheet("Uso de bancos en Electronica B")    
        bancos = ["Banco 1","Banco 2","Banco 3","Banco 4","Banco 5","Banco 6"]
        cantidad_prestamos = []  

        lengDatas = len(bancos)+1

        tab2['B2'].border = thin_border
        tab2['B2'].alignment = Alignment(horizontal='center',vertical='center')
        tab2['B2'] = "Uso de los bancos en Electrónica B"
        tab2.merge_cells('B2:'+get_column_letter(lengDatas)+'2')

        for row in tab2.iter_rows(min_col=2,max_col=lengDatas,min_row=3,max_row=3):
            for (index,cell) in enumerate(row):
                cell.value = bancos[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border
                cantidad_prestamos.append(elemento.find({"Banco":str(index+1)}).count())

        for row in tab2.iter_rows(min_col=2,max_col=lengDatas,min_row=4,max_row=4):
            for (index,cell) in enumerate(row):
                cell.value = cantidad_prestamos[index]
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.border = thin_border

        labels = Reference(tab2,min_col=2,max_col=lengDatas,min_row=3)
        data = Reference(tab2,min_col=2,max_col=lengDatas,min_row=4)      

        chart = BarChart()
        chart.type = "col"
        chart.title = "Cantidad de prestamos para los bancos de Electronica B"
        chart.add_data(data,from_rows=True)
        chart.set_categories(labels)
        chart.legend = None
        chart.varyColors = True
        chart.y_axis.title = "Cantidad de prestamos"  
        chart.x_axis.title = "Bancos de la sala de Electrónica B"
        tab2.add_chart(chart,"B6")

        if 'temporalExcelFilterDate' in mongo.db.list_collection_names():
            temporalExcelFilterDate = mongo.db.temporalExcelFilterDate
            temporalExcelFilterDate.drop()

        return Response(
            save_virtual_workbook(xlsx),
            headers={
            'Content-Disposition': 'attachment; filename=Prueba.xlsx',
            'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

    def post(self, accion):
        mensaje = "error"
        if accion == "registrar":
            mensaje = self.registrar()
        elif accion == "desactivar":
            mensaje = self.desactivar()
        elif accion == "desactivarEquipo":
            mensaje = self.desactivarEquipo()
        elif accion == "activarEquipo":
            mensaje = self.activarEquipo()
        elif accion == "editaruser":
            mensaje = self.editaruser()
        elif accion == "editarequipo":
            mensaje = self.editarequipo()
        elif accion == "registrarEquipo":
            mensaje = self.registrarEquipo()    
        elif accion == "editaruserlab":
            mensaje = self.editaruserlab()
        elif accion == "editpass":
            mensaje = self.editpass()
        elif accion == "editpassadmin":
            mensaje = self.editpassadmin()
        elif accion == "resetpass":
            mensaje = self.resetpass()
        elif accion == "activar":
            mensaje = self.activar()
        elif accion == "buscarUno":
            mensaje = self.buscarUno()
        elif accion == "login":
            mensaje = self.login()   
        elif accion == "borrar":
            mensaje = self.borrarPrueba()
        elif accion == "reserva":
            mensaje = self.reserva()
        elif accion == "buscarreserva":
            respuesta = self.buscarreserva()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "getReservasLaboratorio":
            respuesta = self.getReservasLaboratorio()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "borrarreserva":
            mensaje = self.borrarreserva()
        elif accion == "editarreserva":
            mensaje = self.editarreserva()
        elif accion == "editarAdicional":
            mensaje = self.editarAdicional()
        elif accion == "aprobarreserva":
            mensaje = self.aprobarreserva()
        elif accion == "asistenciareserva":
            mensaje = self.asistenciareserva()            
        elif accion == "consultaeditlabo":
            respuesta = self.consultaeditlabo()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "consultartipo":
            respuesta = self.consultartipo()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "consultaeditadmin":
            respuesta = self.consultaeditadmin()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "consultaTraslados":
            respuesta = self.consultaTraslados()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "consultabanco":
            mensaje = self.consultabanco()
        elif accion == "registrarMant":
            mensaje = self.registrarMant()
        elif accion == "excelTemporalElementos":
            mensaje = self.excelTemporalElementos()
        elif accion == "excelTemporalFilterDate":
            mensaje = self.excelTemporalFilterDate()
        elif accion == "excelTemporalFilterDateBancosSala":
            mensaje = self.excelTemporalFilterDateBancosSala()
        elif accion == "send_mail":
            mensaje = self.send_mail()
        elif accion == "licenciasEstudiantes":
            mensaje = self.licenciasEstudiantes()
        elif accion == "addHorarioAdicional":
            mensaje = self.addHorarioAdicional()
        elif accion == "deleteHorarioAdicional":
            mensaje = self.deleteHorarioAdicional()

        else:
            self.status = 2
            mensaje = "Error en la peticion"
            print(mensaje)     
        return {'status': self.status, 'mensaje': mensaje, 'data': self.datos}

    def get(self, accion):
        mensaje = "error"
        if accion == "consultarTodos":
            respuesta = self.consultarTodos()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "buscarreservalabo":
            respuesta = self.buscarreservalabo()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "buscarhorarios":
            respuesta = self.buscarhorarios()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "logueado":
            mensaje = self.isLoggedIn()
        elif accion == "cerrarSesion":
            mensaje = self.cerrarSesion()
        elif accion == "consultarLabo":
            respuesta = self.consultarLabo()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "consultarEquipo":
            respuesta = self.consultarEquipo()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']
        elif accion == "consultaMantenimiento":
            respuesta = self.consultaMantenimiento()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']           
        elif accion == "getToken":
            mensaje = self.token
        elif accion == "obtenerElementosExcel":
            respuesta = self.obtenerElementosExcel()
            mensaje = respuesta['mensaje']
            self.datos = respuesta['data']     
        elif accion == "altiumLogo":
            return self.altiumLogo()
        elif accion == "psimLogo":
            return self.psimLogo()
        elif accion == "solidWorksLogo":
            return self.solidWorksLogo()
        elif accion == "xirioLogo":
            return self.xirioLogo()
        elif accion == "getHorariosAdicionales":
            respuesta =  self.getHorariosAdicionales()
            mensaje = respuesta ['mensaje']
            self.datos = respuesta['data']

        # Funciones para los informes en excel. El return debe ser de tipo response (Flask)
        elif accion == "createExcel":
            return self.createExcel()          
        elif accion == "excelHorarios_DiasSemanaVsHoras":
            return self.excelHorarios_DiasSemanaVsHoras()      
        elif accion == "excelPrestamos":
            return self.excelPrestamos()
        elif accion == "excelPrestamosBancosSala":
            return self.excelPrestamosBancosSala()
        else:
            self.status = 2
            mensaje = "Error en la peticion"
        return {'status': self.status, 'mensaje': mensaje, 'data': self.datos}


@app.route('/ConsultaSelect2', methods=['POST'])
def consultarTodos():
    usuario = mongo.db.Usuarios
    output = []
    buscar = request.form['q']
    mensaje = "Encontrado"
    for u in usuario.find({"Nombre_usuario": {'$regex': '.*'+buscar+'.*','$options':"i"}}):
        output.append({'id': u['Nombre_usuario'], 'text': u['Correo_usuario'],
                       'estado': u['Código_Estado'], 'identificacion': u['Código_usuario']})
    resp = app.make_response(jsonify(output))
    resp.headers.add('Access-Control-Allow-Origin', '*')
    resp.headers.add('Access-Control-Allow-Credentials', 'true')
    resp.headers.add('Access-Control-Allow-Headers',
                     'Content-Type,Authorization')
    return resp

@app.route('/consAddMantLab', methods=['POST'])
def consultarLabora():
    usuario = mongo.db.Usuarios
    output = []
    buscar = request.form['q']
    mensaje = "Encontrado"
    for u in usuario.find({"$and":[{"Nombre_usuario": {'$regex': '.*'+buscar+'.*','$options':"i"}},{"Tipo":"Laboratorista"}]}):
        output.append({'id': u['Código_usuario'], 'text': u['Código_usuario']+' - '+u['Nombre_usuario'], 'nombre': u['Nombre_usuario']})
    resp = app.make_response(jsonify(output))
    resp.headers.add('Access-Control-Allow-Origin', '*')
    resp.headers.add('Access-Control-Allow-Credentials', 'true')
    resp.headers.add('Access-Control-Allow-Headers',
                     'Content-Type,Authorization')
    return resp

@app.route('/consAddMantPlaca', methods=['POST'])
def consultarPorPlaca():
    elemento = mongo.db.Elementos
    output = []
    buscar = request.form['q']
    mensaje = "Encontrado"
    for u in elemento.find({"$or":[{"Placa": {'$regex': '.*'+buscar+'.*','$options':"i"}},{"Espacio_Fisico": {'$regex': '.*'+buscar+'.*','$options':"i"}},{"Nombre_Del_Equipo": {'$regex': '.*'+buscar+'.*','$options':"i"}}]}):
        output.append({'id': u['Placa'], 'text': u['Placa']+'-'+u['Nombre_Del_Equipo'],'numInterno':u['Codigo_Interno'],'nomEqu':u['Nombre_Del_Equipo'],'sala':u['Espacio_Fisico']})
    resp = app.make_response(jsonify(output))
    resp.headers.add('Access-Control-Allow-Origin', '*')
    resp.headers.add('Access-Control-Allow-Credentials', 'true')
    resp.headers.add('Access-Control-Allow-Headers',
                     'Content-Type,Authorization')
    return resp


@app.route('/login', methods=['POST'])
def saludo():
    user = mongo.db.Usuarios
    usuario = request.json['usuario']
    contrasena = request.json['contrasena']
    pregunta = user.find_one(
        {"Código_usuario": usuario, "Contraseña": contrasena})
    token = ''
    if pregunta:
        session['tipoUsuario'] = "Admin"
        tokenGen = urandom(64)
        session['token'] = b64encode(tokenGen).decode('utf-8')
        token = session['token']
        mensaje = "1"
        status = 1
    else:
        mensaje = "0"
        status = 1
    resp = app.make_response(
        jsonify({'status': status, 'mensaje': mensaje, 'data': {'token': token}}))
    resp.set_cookie('Autorizacion', 'si')
    resp.headers.add('Access-Control-Allow-Origin', '*')
    resp.headers.add('Access-Control-Allow-Credentials', 'true')
    resp.headers.add('Access-Control-Allow-Methods',
                     'GET,PUT,POST,DELETE,OPTIONS')
    resp.headers.add('Access-Control-Allow-Headers',
                     'Content-Type,Authorization')
    return resp


api.add_resource(Usuario, '/Usuario/<string:accion>')  # Route_1

if __name__ == '__main__':
    app.secret_key = 'ssssshhhhh'
    app.run(port='7001', ssl_context='adhoc')


# @app.route('/send_mail')
# def send_mail():
#     print("ENTRO A MAIL WIIII")
#     msg = mail.send_message(
#         'Send Mail tutorial!',
#         sender='adminsalas@udistrital.edu.co',
#         recipients=['rjchiaj@correo.udistrital.edu.co'],
#         body="Prueba del correo."
#     )
#     return 'Mail sent'

